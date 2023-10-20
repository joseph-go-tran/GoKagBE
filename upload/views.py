import datetime

import numpy as np
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.decorators import action, parser_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

from api.exceptions import (AnswerNotEnoughException,
                            FileUploadRequireExcepion,
                            QuestionnaireIdRequireExcepion)
from api.utils import ObjectResponse, StatusResponse, try_except_wrapper
from datasets.serializer import AnswerSerializer
from questionnaire.models import Questionnaire
from questionnaire.serializer import QuestionnaireSerializer
from questionnaire.views.views_question import QuestionViewSet


class UploadViewSet(QuestionViewSet):
    @action(
        methods=["POST"],
        url_path="question",
        url_name="upload_question",
        detail=False,
    )
    @parser_classes([MultiPartParser, FormParser])
    @try_except_wrapper
    def create_question(self, request):
        """
        CREATE QUESTION FOR QUESTIONNAIRE FROM FILE
        """
        file = request.FILES.get("file")
        questionnaire = request.POST.get("questionnaire")

        if file is None:
            raise FileUploadRequireExcepion
        if questionnaire is None:
            raise QuestionnaireIdRequireExcepion

        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active

        self.remove_empty_rows_and_columns(sheet)

        self.handle_create_question_for_questionnaire(request, sheet)

        return Response(
            ObjectResponse(
                StatusResponse.STATUS_SUCCESS,
                "Create questions for questionnaire successfully.",
                "",
            ).get_json(),
            status=status.HTTP_200_OK,
        )

    def handle_create_question_for_questionnaire(self, request, sheet):
        for idx, col in enumerate(sheet.iter_cols(values_only=True), start=1):
            if isinstance(col[1], datetime.datetime):
                # Handle number datetime type
                continue

            # print(idx)
            self.handle_column_question(request, col, sequence=idx)
        return True

    def handle_column_question(self, request, data, sequence):
        questionnaire = request.POST.get("questionnaire")
        data_clean = tuple(item for item in data if item is not None)

        required = True
        if len(data) != len(data_clean):
            required = False

        # Check number
        data_clean = [str(x) if isinstance(x, int) else x for x in data_clean]

        data_process = [item.split(", ") for item in data_clean[1:]]
        flat_data = []
        for sublist in data_process:
            flat_data.extend(sublist)

        if self.is_number_input(flat_data):
            type_question = self.check_question_number_type(
                data_process, flat_data
            )
        else:
            type_question = self.check_question_type(data_clean[1:])

        # print(type_question)
        if type_question == "InputType":
            obj = self.process_InputType_object(
                data_clean, questionnaire, sequence, required
            )
        elif type_question == "SelectType":
            obj = self.process_SelectType_object(
                data_clean, questionnaire, sequence, required, multi=False
            )
        elif type_question == "SelectType-multi":
            obj = self.process_SelectType_object(
                data_clean, questionnaire, sequence, required, multi=True
            )
        self.handle_create(request, obj)

    def process_InputType_object(
        self, data, questionnaire, sequence, required
    ):
        return {
            "questionnaire": questionnaire,
            "type": "InputType",
            "label": data[0],
            "sequence": sequence,
            "question_detail": {"placeholder": "", "required": required},
        }

    def process_SelectType_object(
        self, data, questionnaire, sequence, required, multi
    ):
        options_list = (
            [
                item
                for sublist in [item.split(", ") for item in data]
                for item in sublist
            ]
            if multi
            else data
        )

        options_set = set(options_list[1:])
        sorted_data = options_set

        if self.is_number_input(options_list[1:]):
            sorted_data = sorted(options_set)
        options = [{"value": item} for item in sorted_data]

        return {
            "questionnaire": questionnaire,
            "type": "SelectType",
            "label": data[0],
            "sequence": sequence,
            "question_detail": {
                "multiselect": multi,
                "html_select": False,
                "required": required,
                "options": options,
            },
        }

    @action(
        methods=["POST"],
        url_path="answer",
        url_name="upload_answer",
        detail=False,
    )
    @parser_classes([MultiPartParser, FormParser])
    @try_except_wrapper
    def create_answer(self, request):
        """
        CREATE ANSWER LIST FROM EXCELL FILE
        """
        file = request.FILES.get("file")
        questionnaire = request.POST.get("questionnaire")

        if file is None:
            raise FileUploadRequireExcepion
        if questionnaire is None:
            raise QuestionnaireIdRequireExcepion

        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active

        self.remove_empty_rows_and_columns(sheet)

        self.handle_create_answer_for_dataset(request, sheet)

        return Response(
            ObjectResponse(
                StatusResponse.STATUS_SUCCESS,
                "Create answer successfully.",
                "",
            ).get_json(),
            status=status.HTTP_200_OK,
        )

    @action(
        methods=["POST"],
        url_path="datasets",
        url_name="upload_datasets",
        detail=False,
    )
    @parser_classes([MultiPartParser, FormParser])
    @try_except_wrapper
    def create_datasets(self, request):
        """
        CREATE DATASETS FROM EXCELL FILE
        """
        file = request.FILES.get("file")
        questionnaire = request.POST.get("questionnaire")

        if file is None:
            raise FileUploadRequireExcepion
        if questionnaire is None:
            raise QuestionnaireIdRequireExcepion

        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active

        self.remove_empty_rows_and_columns(sheet)

        self.handle_create_question_for_questionnaire(request, sheet)
        self.handle_create_answer_for_dataset(request, sheet)

        return Response(
            ObjectResponse(
                StatusResponse.STATUS_SUCCESS,
                "Create answer successfully.",
                "",
            ).get_json(),
            status=status.HTTP_200_OK,
        )

    def handle_create_answer_for_dataset(self, request, sheet):
        questionnaire = request.POST.get("questionnaire")

        first_column = None
        for col in sheet.iter_cols(values_only=True):
            first_column = col
            break
        if isinstance(first_column[1], datetime.datetime):
            sheet.delete_cols(1)

        num_columns = sheet.max_column
        # print(num_columns)

        # Get questionnaire from DB
        instance = get_object_or_404(Questionnaire, pk=questionnaire)
        serializer = QuestionnaireSerializer(instance)
        questionnaire_data = serializer.data

        # print(len(questionnaire_data["questions"]))
        if num_columns != len(questionnaire_data["questions"]):
            raise AnswerNotEnoughException

        # Check the data type of each column that does not match the question
        for idx, col in enumerate(sheet.iter_cols(values_only=True), start=1):
            data_clean = tuple(item for item in col if item is not None)
            data_clean = [
                str(x) if isinstance(x, int) else x for x in data_clean
            ]
            data_process = [item.split(", ") for item in data_clean[1:]]
            flat_data = []
            for sublist in data_process:
                flat_data.extend(sublist)

            if self.is_number_input(flat_data):
                type_question = self.check_question_number_type(
                    data_process, flat_data
                )
            else:
                type_question = self.check_question_type(data_clean[1:])
            data_type = type_question.split("-")[0]
            question_type = questionnaire_data["questions"][idx - 1]["type"]

            if data_type != question_type:
                return Response(
                    ObjectResponse(
                        StatusResponse.STATUS_SUCCESS,
                        "Data type is not comfortable.",
                        "Columns: " + col[0],
                    ).get_json(),
                    status=status.HTTP_200_OK,
                )
        for code, row in enumerate(sheet.iter_rows(values_only=True), start=0):
            if isinstance(col[1], datetime.datetime):
                # Handle number datetime type
                continue

            if code == 0:  # Skip the first row
                continue

            isValid = True
            data_save = []
            for idx, item in enumerate(row, start=0):
                question_type = questionnaire_data["questions"][idx]["type"]
                # Check question is required but item is None
                isValid = isValid and self.is_valid_answer_item_required(
                    item, questionnaire_data["questions"][idx]
                )
                if item is not None and question_type == "SelectType":
                    isValid = isValid and self.is_valid_answer_item_SelectType(
                        item, questionnaire_data["questions"][idx]
                    )
                if isValid is False:
                    break
                data_save.append(
                    {
                        "questionnaire": questionnaire,
                        "question_key": questionnaire_data["questions"][idx][
                            "key"
                        ],
                        "value": item,
                        "code": code,
                    }
                )

            if isValid:
                for item in data_save:
                    answer_serializer = AnswerSerializer(data=item)
                    answer_serializer.is_valid(raise_exception=True)
                    answer_serializer.save(answer_by=request.user)

    def is_valid_answer_item_SelectType(self, item, question):
        item = str(item)
        question_detail = question["question_detail"]
        if question_detail["multiselect"]:  # multiselect
            options_list = question_detail["options"]
            item_list = item.split(", ")

            for e in options_list:
                if e["value"] in item_list:
                    item_list.remove(e["value"])

            if (
                question_detail["other_field"] is False and len(item_list) > 0
            ) or (question_detail["other_field"] and len(item_list) > 1):
                return False
        elif (  # singleselect
            question_detail["multiselect"] is False
            and question_detail["other_field"] is False
            and not any(e["value"] == item for e in question_detail["options"])
        ):
            return False

        return True

    def is_valid_answer_item_required(self, item, question):
        question_detail = question["question_detail"]
        if item is None and question_detail["required"]:
            return False
        return True

    def check_question_type(self, data):
        result = self.calculate_similarity(data)
        # print(result)
        if result["mean_similar"] >= 0.25:
            return "SelectType"
        else:
            if result["similar_ratio"] >= 0.4:
                return "SelectType-multi"
            else:
                return "InputType"

    def calculate_similarity(self, strings):
        """
        Calculates similarity metrics based on TF-IDF
        and cosine similarity for a list of strings.

        TF-IDF (Term Frequency-Inverse Document Frequency) is a technique
        used in NLP and information retrieval.
        It measures the importance of a word within a document
        relative to a larger set of documents.

        Cosine Similarity is a measure of similarity between
        two non-zero vectors in an inner product space.
        """
        vectorizer = TfidfVectorizer()

        commas_count = [string.count(", ") for string in strings]
        commas_count = np.sum(commas_count)

        tfidf_matrix = vectorizer.fit_transform(strings)

        similarity_matrix = cosine_similarity(tfidf_matrix)

        num_ones = np.count_nonzero(
            np.isclose(similarity_matrix, 1, rtol=1e-2, atol=1e-2)
        )
        num_elements = similarity_matrix.shape[0] * similarity_matrix.shape[1]

        return {
            "similar_ratio": np.mean(similarity_matrix),
            "mean_similar": num_ones / num_elements,
            "mean_commas": commas_count / len(strings),
        }

    def is_convertible_to_int(self, s):
        """
        Check a string can be convert to int
        """
        try:
            int(s)
            return True
        except ValueError:
            return False

    def is_number_input(self, data):
        """
        Check data is a number type
        """
        for item in data:
            if self.is_convertible_to_int(item) is False:
                return False
        return True

    def check_question_number_type(self, data, flat_data):
        for item in data:
            if len(item) > 1:
                return "SelectType-multi"

        unique_values, counts = np.unique(flat_data, return_counts=True)
        total_frequency = np.sum(counts)

        # Tính trung bình tần số xuất hiện
        mean_frequency = total_frequency / len(counts)
        print(mean_frequency)
        if mean_frequency <= 1.5:
            return "InputType"
        return "SelectType"

    def remove_empty_rows_and_columns(self, sheet):
        rows_to_remove = []
        cols_to_remove = []

        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=1, values_only=True), start=1
        ):
            if all(cell is None for cell in row):
                rows_to_remove.append(row_idx)

        for col_idx, col in enumerate(
            sheet.iter_cols(min_col=1, values_only=True), start=1
        ):
            if all(cell is None for cell in col):
                cols_to_remove.append(col_idx)

        for row_idx in reversed(rows_to_remove):
            sheet.delete_rows(row_idx)

        for col_idx in reversed(cols_to_remove):
            sheet.delete_cols(col_idx)
